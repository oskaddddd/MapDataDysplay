import openpyxl
import json

seperatorCharacter = ';'
dat = []
data = []
f =  openpyxl.load_workbook("ExamleExcel.xlsx")
f = f.active
    
for i in range(0, f.max_row):
    t = {}
    do = True
    for i1, col in enumerate(f.iter_cols(1, f.max_column)):
        if i1 == 0:
            gps = []
            if ', ' in col[i].value:
                gps = col[i].value.split()
                gps[0] = gps[0][:len(gps[0])-1]
                t['GPS'] = gps
                #print(1)
            elif ','in col[i].value and ' ' not in col[i].value:
                gps = [col[i].value[:col[i].value.index(',')], col[i].value[col[i].value.index(',')+1:]]
                t['GPS'] = gps
                #print(2)
            elif ',' not in col[i].value and ' ' in col[i].value:
                gps = col[i].value.split()
                t['GPS'] = gps
                #print(3)
            else:
                do = False
            if do == True:
                try:
                    print(gps)
                    float(gps[0])
                    float(gps[1])
                except Exception as e:
                    do = False
        elif i1 == 1: 
            t["Value"] = col[i].value
            try: float(col[i].value)
            except: 
                do = False
    print(t, f'Include = {do}')
    if do == True:
        dat.append(t)
print(dat)

with open(f"data.json", 'w') as f:
    json.dump(dat, f, indent=1)