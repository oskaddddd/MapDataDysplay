import openpyxl
import json

seperatorCharacter = ';'
dat = []
f =  openpyxl.load_workbook("ExamleExcel.xlsx")
f = f.active
    
for i in range(0, f.max_row):
    t = []
    for col in f.iter_cols(1, f.max_column):
        t.append(col[i].value)
    dat.append(t)
print(dat)

try: int(dat[0][1])
except: dat.pop(0)

print(dat)
for i, x in enumerate(dat):
    t = []
    if ', ' in x[0]:
        t = x[0].split()
        t[0] = t[0][:len(t[0])-1]
        print(1)
    elif ','in x[0] and ' ' not in x[0]:
        t = [x[0][:x[0].index(',')], x[0][x[0].index(',')+1:]]
        print(2)
    elif ',' not in x[0] and ' ' in x[0]:
        t = x[0].split()
        print(3)
    dat[i][0] = t
    print(t)

with open(f"data.json", 'w') as f:
    json.dump(dat, f, indent=1)
        

        
        